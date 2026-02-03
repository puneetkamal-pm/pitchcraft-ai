"""
Professional DCF Model Generator
Creates IB-quality DCF models in Excel with all the bells and whistles
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule
from openpyxl.chart import BarChart, Reference
from typing import Optional
from core.question_generator import DCFAssumptions


class ProfessionalDCFModel:
    """Generate IB-quality DCF model with professional features"""

    TAB_ORDER = [
        "Cover",
        "Contents",
        "Inputs_Index",
        "Key_Assumptions",
        "Historical_IS",
        "Historical_BS",
        "Historical_CF",
        "Revenue_Build",
        "COGS_Gross_Margin",
        "Opex",
        "EBITDA_Bridge",
        "D&A",
        "Capex",
        "Working_Capital",
        "Other_Operating",
        "Taxes",
        "Unlevered_FCF",
        "Debt_Schedule",
        "Interest_Expense",
        "Share_Count",
        "WACC",
        "DCF_Valuation",
        "Terminal_Value",
        "EV_Equity_Bridge",
        "Sensitivity",
        "Scenario_Manager",
        "KPI_Dashboard",
        "Trading_Comps",
        "Transactions_Comps",
        "Charts_Checks",
    ]

    # Style definitions - IB standard colors
    STYLES = {
        'input_font': Font(color="0000FF"),  # Blue for inputs
        'formula_font': Font(color="000000"),  # Black for formulas
        'link_font': Font(color="008000"),  # Green for cross-sheet links
        'header_font': Font(bold=True, size=11),
        'title_font': Font(bold=True, size=14),
        'section_font': Font(bold=True, size=11, color="FFFFFF"),
        'section_fill': PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid"),
        'input_fill': PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        'output_fill': PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),
        'error_fill': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        ),
        'border_bottom': Border(bottom=Side(style='thin')),
    }

    def __init__(self, assumptions: DCFAssumptions):
        self.a = assumptions
        self.wb = Workbook()
        self._cell_refs = {}  # Store cell references for formulas

        # Computed values for compatibility with model
        # NWC from cash conversion cycle (DSO + DIO - DPO)
        ccc = (getattr(assumptions, 'days_sales_outstanding', 45) +
               getattr(assumptions, 'days_inventory_outstanding', 15) -
               getattr(assumptions, 'days_payables_outstanding', 35))
        self._nwc_pct_revenue = ccc / 365

        # WACC buildup values
        self._unlevered_beta = getattr(assumptions, 'unlevered_beta', 1.0)
        tax_rate = assumptions.tax_rate
        target_de = getattr(assumptions, 'target_debt_to_equity', 0.25)
        self._levered_beta = self._unlevered_beta * (1 + (1 - tax_rate) * target_de)
        self._cost_of_debt = getattr(assumptions, 'pre_tax_cost_of_debt', 0.06)
        self._debt_to_capital = target_de / (1 + target_de)

        # Size and company-specific premiums
        self._size_premium = getattr(assumptions, 'size_premium', 0)
        self._company_specific_risk = getattr(assumptions, 'company_specific_risk', 0)

        # Net debt computed from total_debt and cash
        self._net_debt = getattr(assumptions, 'total_debt', 0) - getattr(assumptions, 'cash', 0)

    def _setup_sheet(self, ws, title: str):
        """Apply standard formatting to a worksheet"""
        ws.title = title
        ws.column_dimensions['A'].width = 35
        for i in range(2, 15):
            ws.column_dimensions[get_column_letter(i)].width = 14

    def _add_title(self, ws, row: int, text: str) -> int:
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = self.STYLES['title_font']
        return row + 2

    def _add_section_header(self, ws, row: int, text: str, span: int = 8) -> int:
        """Add a section header row"""
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = self.STYLES['section_font']
        cell.fill = self.STYLES['section_fill']
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        return row + 1

    def _add_input_cell(self, ws, row: int, col: int, value, fmt: str = None, name: str = None):
        """Add an input cell with blue font and yellow background"""
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = self.STYLES['input_font']
        cell.fill = self.STYLES['input_fill']
        cell.border = self.STYLES['border']
        if fmt:
            cell.number_format = fmt
        if name:
            self._cell_refs[name] = f"{get_column_letter(col)}{row}"
        return cell

    def _add_formula_cell(self, ws, row: int, col: int, formula: str, fmt: str = None, bold: bool = False):
        """Add a formula cell"""
        cell = ws.cell(row=row, column=col, value=formula)
        cell.font = Font(bold=bold) if bold else self.STYLES['formula_font']
        if fmt:
            cell.number_format = fmt
        return cell

    def _ref(self, name: str) -> str:
        """Get cell reference by name"""
        return self._cell_refs.get(name, "")

    def build_cover_sheet(self):
        """Build the Cover tab"""
        ws = self.wb.active
        self._setup_sheet(ws, "Cover")
        row = 1
        row = self._add_title(ws, row, f"DCF Valuation Model - {self.a.company_name} ({self.a.ticker})")
        ws.cell(row=row, column=1, value="Prepared For: Investment Banking Associates")
        row += 1
        ws.cell(row=row, column=1, value="Prepared By: PitchCraftAI")
        row += 1
        ws.cell(row=row, column=1, value="As of: February 3, 2026")
        return ws

    def build_contents_sheet(self):
        """Build the Contents tab with a table of contents"""
        ws = self.wb.create_sheet("Contents")
        self._setup_sheet(ws, "Contents")
        row = 1
        row = self._add_title(ws, row, "Model Contents")

        ws.cell(row=row, column=1, value="Tab")
        ws.cell(row=row, column=2, value="Purpose")
        ws.cell(row=row, column=1).font = self.STYLES['header_font']
        ws.cell(row=row, column=2).font = self.STYLES['header_font']
        row += 1

        tabs = [
            ("Cover", "Model cover"),
            ("Contents", "Table of contents"),
            ("Inputs_Index", "Key inputs and controls"),
            ("Key_Assumptions", "Core model assumptions"),
            ("Historical_IS", "Historical income statement"),
            ("Historical_BS", "Historical balance sheet"),
            ("Historical_CF", "Historical cash flow"),
            ("Revenue_Build", "Revenue forecast build"),
            ("COGS_Gross_Margin", "Gross margin / COGS build"),
            ("Opex", "Operating expenses"),
            ("EBITDA_Bridge", "EBITDA derivation"),
            ("D&A", "Depreciation & amortization"),
            ("Capex", "Capital expenditures"),
            ("Working_Capital", "Net working capital build"),
            ("Other_Operating", "Other operating items"),
            ("Taxes", "Tax schedule"),
            ("Unlevered_FCF", "Unlevered free cash flow"),
            ("Debt_Schedule", "Debt and amortization"),
            ("Interest_Expense", "Interest expense schedule"),
            ("Share_Count", "Basic/diluted shares"),
            ("WACC", "WACC build"),
            ("DCF_Valuation", "DCF valuation"),
            ("Terminal_Value", "Terminal value methods"),
            ("EV_Equity_Bridge", "EV to equity bridge"),
            ("Sensitivity", "Sensitivity tables"),
            ("Scenario_Manager", "Scenario controls"),
            ("KPI_Dashboard", "Key KPIs"),
            ("Trading_Comps", "Trading comps"),
            ("Transactions_Comps", "Transactions comps"),
            ("Charts_Checks", "Charts and error checks"),
        ]

        for tab, purpose in tabs:
            ws.cell(row=row, column=1, value=tab)
            ws.cell(row=row, column=2, value=purpose)
            row += 1

        return ws

    def build_inputs_index_sheet(self):
        """Build an index of key inputs"""
        ws = self.wb.create_sheet("Inputs_Index")
        self._setup_sheet(ws, "Inputs_Index")
        row = 1
        row = self._add_title(ws, row, "Inputs Index")

        ws.cell(row=row, column=1, value="Input")
        ws.cell(row=row, column=2, value="Value")
        ws.cell(row=row, column=1).font = self.STYLES['header_font']
        ws.cell(row=row, column=2).font = self.STYLES['header_font']
        row += 1

        inputs = [
            ("Base Revenue ($M)", "Key_Assumptions", "base_revenue"),
            ("EBITDA Margin", "Key_Assumptions", "ebitda_margin"),
            ("D&A (% Rev)", "Key_Assumptions", "da_pct"),
            ("Capex (% Rev)", "Key_Assumptions", "capex_pct"),
            ("NWC (% Rev)", "Key_Assumptions", "nwc_pct"),
            ("Tax Rate", "Key_Assumptions", "tax_rate"),
            ("WACC", "WACC", "wacc"),
            ("Terminal Growth", "Key_Assumptions", "term_growth"),
            ("Exit EV/EBITDA", "Key_Assumptions", "exit_mult"),
            ("Net Debt ($M)", "Key_Assumptions", "net_debt"),
            ("Shares (M)", "Key_Assumptions", "shares"),
        ]

        for label, sheet, ref in inputs:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=f"={sheet}!{self._ref(ref)}" if self._ref(ref) else "—")
            row += 1

        return ws

    def build_key_assumptions_sheet(self):
        """Build the Key Assumptions tab"""
        ws = self.wb.create_sheet("Key_Assumptions")
        self._setup_sheet(ws, "Key_Assumptions")

        row = 1
        row = self._add_title(ws, row, f"Key Assumptions - {self.a.company_name} ({self.a.ticker})")

        # Company Info
        row = self._add_section_header(ws, row, "COMPANY INFORMATION", 4)
        ws.cell(row=row, column=1, value="Company Name")
        ws.cell(row=row, column=2, value=self.a.company_name)
        row += 1
        ws.cell(row=row, column=1, value="Ticker")
        ws.cell(row=row, column=2, value=self.a.ticker)
        row += 2

        # Revenue Assumptions
        row = self._add_section_header(ws, row, "REVENUE ASSUMPTIONS", 4)
        ws.cell(row=row, column=1, value="Base Year Revenue ($M)")
        self._add_input_cell(ws, row, 2, self.a.base_revenue, '#,##0', 'base_revenue')
        row += 1

        growth_rates = [
            ("Year 1 Growth", self.a.revenue_growth_y1, 'growth_y1'),
            ("Year 2 Growth", self.a.revenue_growth_y2, 'growth_y2'),
            ("Year 3 Growth", self.a.revenue_growth_y3, 'growth_y3'),
            ("Year 4 Growth", self.a.revenue_growth_y4, 'growth_y4'),
            ("Year 5 Growth", self.a.revenue_growth_y5, 'growth_y5'),
        ]
        for label, value, name in growth_rates:
            ws.cell(row=row, column=1, value=label)
            self._add_input_cell(ws, row, 2, value, '0.0%', name)
            row += 1
        row += 1

        # Operating Assumptions
        row = self._add_section_header(ws, row, "OPERATING ASSUMPTIONS", 4)
        op_assumptions = [
            ("EBITDA Margin", self.a.ebitda_margin, '0.0%', 'ebitda_margin'),
            ("D&A (% of Revenue)", self.a.da_pct_revenue, '0.0%', 'da_pct'),
            ("CapEx (% of Revenue)", self.a.capex_pct_revenue, '0.0%', 'capex_pct'),
            ("NWC (% of Revenue)", self._nwc_pct_revenue, '0.0%', 'nwc_pct'),
            ("Tax Rate", self.a.tax_rate, '0.0%', 'tax_rate'),
        ]
        for label, value, fmt, name in op_assumptions:
            ws.cell(row=row, column=1, value=label)
            self._add_input_cell(ws, row, 2, value, fmt, name)
            row += 1
        row += 1

        # WACC Buildup
        row = self._add_section_header(ws, row, "WACC BUILDUP", 4)
        wacc_inputs = [
            ("Risk-Free Rate", self.a.risk_free_rate, '0.00%', 'rf_rate'),
            ("Equity Risk Premium", self.a.equity_risk_premium, '0.00%', 'erp'),
            ("Unlevered Beta", self._unlevered_beta, '0.00', 'unlevered_beta'),
            ("Levered Beta", self._levered_beta, '0.00', 'beta'),
            ("Size Premium", self._size_premium, '0.00%', 'size_prem'),
            ("Company Risk Premium", self._company_specific_risk, '0.00%', 'co_risk'),
            ("Pre-Tax Cost of Debt", self._cost_of_debt, '0.00%', 'cost_debt'),
            ("Debt / Total Capital", self._debt_to_capital, '0.0%', 'debt_cap'),
        ]
        for label, value, fmt, name in wacc_inputs:
            ws.cell(row=row, column=1, value=label)
            self._add_input_cell(ws, row, 2, value, fmt, name)
            row += 1

        # Calculated WACC
        row += 1
        ws.cell(row=row, column=1, value="Cost of Equity (CAPM + Adj)")
        cost_equity_formula = f"={self._ref('rf_rate')}+{self._ref('beta')}*{self._ref('erp')}+{self._ref('size_prem')}+{self._ref('co_risk')}"
        self._add_formula_cell(ws, row, 2, cost_equity_formula, '0.00%')
        self._cell_refs['cost_equity'] = f"B{row}"
        row += 1

        ws.cell(row=row, column=1, value="After-Tax Cost of Debt")
        atax_debt_formula = f"={self._ref('cost_debt')}*(1-{self._ref('tax_rate')})"
        self._add_formula_cell(ws, row, 2, atax_debt_formula, '0.00%')
        self._cell_refs['atax_cost_debt'] = f"B{row}"
        row += 1

        ws.cell(row=row, column=1, value="WACC")
        ws.cell(row=row, column=1).font = Font(bold=True)
        wacc_formula = f"={self._ref('cost_equity')}*(1-{self._ref('debt_cap')})+{self._ref('atax_cost_debt')}*{self._ref('debt_cap')}"
        cell = self._add_formula_cell(ws, row, 2, wacc_formula, '0.00%', bold=True)
        cell.fill = self.STYLES['output_fill']
        cell.border = self.STYLES['border']
        self._cell_refs['wacc'] = f"B{row}"
        row += 2

        # Terminal Value Assumptions
        row = self._add_section_header(ws, row, "TERMINAL VALUE ASSUMPTIONS", 4)
        ws.cell(row=row, column=1, value="Perpetuity Growth Rate")
        self._add_input_cell(ws, row, 2, self.a.terminal_growth, '0.00%', 'term_growth')
        row += 1
        ws.cell(row=row, column=1, value="Exit EV/EBITDA Multiple")
        self._add_input_cell(ws, row, 2, self.a.exit_ebitda_multiple, '0.0x', 'exit_mult')
        row += 2

        # Capital Structure
        row = self._add_section_header(ws, row, "CAPITAL STRUCTURE", 4)
        ws.cell(row=row, column=1, value="Net Debt ($M)")
        self._add_input_cell(ws, row, 2, self._net_debt, '#,##0', 'net_debt')
        row += 1
        ws.cell(row=row, column=1, value="Shares Outstanding (M)")
        self._add_input_cell(ws, row, 2, self.a.shares_outstanding, '#,##0.0', 'shares')
        row += 2

        # Model Settings
        row = self._add_section_header(ws, row, "MODEL SETTINGS", 4)
        ws.cell(row=row, column=1, value="Use Mid-Year Convention")
        self._add_input_cell(ws, row, 2, 1 if self.a.use_mid_year_convention else 0, '0', 'mid_year')
        ws.cell(row=row, column=3, value="(1=Yes, 0=No)")

        return ws

    def build_historical_sheets(self):
        """Build placeholder historical financials tabs"""
        for name in ["Historical_IS", "Historical_BS", "Historical_CF"]:
            ws = self.wb.create_sheet(name)
            self._setup_sheet(ws, name)
            row = 1
            row = self._add_title(ws, row, f"{name.replace('_', ' ')}")
            ws.cell(row=row, column=1, value="(Populate with historical financials)")
        return True

    def build_revenue_build_sheet(self):
        ws = self.wb.create_sheet("Revenue_Build")
        self._setup_sheet(ws, "Revenue_Build")
        row = 1
        row = self._add_title(ws, row, "Revenue Build")

        years = ["Base", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
        for i, year in enumerate(years):
            cell = ws.cell(row=row, column=i+2, value=year)
            cell.font = self.STYLES['header_font']
            cell.alignment = Alignment(horizontal='center')
        row += 1

        ws.cell(row=row, column=1, value="Revenue Growth")
        ws.cell(row=row, column=2, value="—")
        growth_refs = ['growth_y1', 'growth_y2', 'growth_y3', 'growth_y4', 'growth_y5']
        for i, ref in enumerate(growth_refs):
            cell = ws.cell(row=row, column=i+3, value=f"=Key_Assumptions!{self._ref(ref)}")
            cell.font = self.STYLES['link_font']
            cell.number_format = '0.0%'
        row += 1

        ws.cell(row=row, column=1, value="Revenue")
        ws.cell(row=row, column=2, value=f"=Key_Assumptions!{self._ref('base_revenue')}")
        ws.cell(row=row, column=2).number_format = '#,##0'
        for i in range(5):
            col = i + 3
            prev_col = get_column_letter(col - 1)
            curr_col = get_column_letter(col)
            formula = f"={prev_col}{row}*(1+{curr_col}{row-1})"
            self._add_formula_cell(ws, row, col, formula, '#,##0')
        return ws

    def build_cogs_gross_margin_sheet(self):
        ws = self.wb.create_sheet("COGS_Gross_Margin")
        self._setup_sheet(ws, "COGS_Gross_Margin")
        row = 1
        row = self._add_title(ws, row, "COGS & Gross Margin")
        ws.cell(row=row, column=1, value="(Placeholder for COGS and gross margin build)")
        return ws

    def build_opex_sheet(self):
        ws = self.wb.create_sheet("Opex")
        self._setup_sheet(ws, "Opex")
        row = 1
        row = self._add_title(ws, row, "Operating Expenses")
        ws.cell(row=row, column=1, value="(Placeholder for opex build)")
        return ws

    def build_ebitda_bridge_sheet(self):
        ws = self.wb.create_sheet("EBITDA_Bridge")
        self._setup_sheet(ws, "EBITDA_Bridge")
        row = 1
        row = self._add_title(ws, row, "EBITDA Bridge")

        years = ["Base", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
        for i, year in enumerate(years):
            cell = ws.cell(row=row, column=i+2, value=year)
            cell.font = self.STYLES['header_font']
            cell.alignment = Alignment(horizontal='center')
        row += 1

        ws.cell(row=row, column=1, value="Revenue")
        for i in range(6):
            col = i + 2
            ws.cell(row=row, column=col, value=f"=Revenue_Build!{get_column_letter(col)}5").number_format = '#,##0'
        row += 1

        ws.cell(row=row, column=1, value="EBITDA Margin")
        for i in range(6):
            col = i + 2
            cell = ws.cell(row=row, column=col, value=f"=Key_Assumptions!{self._ref('ebitda_margin')}")
            cell.number_format = '0.0%'
        row += 1

        ws.cell(row=row, column=1, value="EBITDA")
        ws.cell(row=row, column=1).font = Font(bold=True)
        for i in range(6):
            col = i + 2
            formula = f"={get_column_letter(col)}{row-2}*{get_column_letter(col)}{row-1}"
            self._add_formula_cell(ws, row, col, formula, '#,##0', bold=True)
        return ws

    def build_da_sheet(self):
        ws = self.wb.create_sheet("D&A")
        self._setup_sheet(ws, "D&A")
        row = 1
        row = self._add_title(ws, row, "Depreciation & Amortization")
        years = ["Base", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
        for i, year in enumerate(years):
            cell = ws.cell(row=row, column=i+2, value=year)
            cell.font = self.STYLES['header_font']
            cell.alignment = Alignment(horizontal='center')
        row += 1
        ws.cell(row=row, column=1, value="D&A (% Rev)")
        for i in range(6):
            col = i + 2
            cell = ws.cell(row=row, column=col, value=f"=Key_Assumptions!{self._ref('da_pct')}")
            cell.number_format = '0.0%'
        row += 1
        ws.cell(row=row, column=1, value="D&A")
        for i in range(6):
            col = i + 2
            formula = f"=-Revenue_Build!{get_column_letter(col)}5*{get_column_letter(col)}{row-1}"
            self._add_formula_cell(ws, row, col, formula, '#,##0')
        return ws

    def build_capex_sheet(self):
        ws = self.wb.create_sheet("Capex")
        self._setup_sheet(ws, "Capex")
        row = 1
        row = self._add_title(ws, row, "Capital Expenditures")
        years = ["Base", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
        for i, year in enumerate(years):
            cell = ws.cell(row=row, column=i+2, value=year)
            cell.font = self.STYLES['header_font']
            cell.alignment = Alignment(horizontal='center')
        row += 1
        ws.cell(row=row, column=1, value="Capex (% Rev)")
        for i in range(6):
            col = i + 2
            cell = ws.cell(row=row, column=col, value=f"=Key_Assumptions!{self._ref('capex_pct')}")
            cell.number_format = '0.0%'
        row += 1
        ws.cell(row=row, column=1, value="Capex")
        for i in range(6):
            col = i + 2
            formula = f"=-Revenue_Build!{get_column_letter(col)}5*{get_column_letter(col)}{row-1}"
            self._add_formula_cell(ws, row, col, formula, '#,##0')
        return ws

    def build_working_capital_sheet(self):
        ws = self.wb.create_sheet("Working_Capital")
        self._setup_sheet(ws, "Working_Capital")
        row = 1
        row = self._add_title(ws, row, "Working Capital")
        years = ["Base", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
        for i, year in enumerate(years):
            cell = ws.cell(row=row, column=i+2, value=year)
            cell.font = self.STYLES['header_font']
            cell.alignment = Alignment(horizontal='center')
        row += 1
        ws.cell(row=row, column=1, value="NWC (% Rev)")
        for i in range(6):
            col = i + 2
            cell = ws.cell(row=row, column=col, value=f"=Key_Assumptions!{self._ref('nwc_pct')}")
            cell.number_format = '0.0%'
        row += 1
        ws.cell(row=row, column=1, value="Change in NWC")
        ws.cell(row=row, column=2, value=0).number_format = '#,##0'
        for i in range(5):
            col = i + 3
            prev_col = get_column_letter(col - 1)
            curr_col = get_column_letter(col)
            formula = f"=-({curr_col}5-{prev_col}5)*{curr_col}{row-1}"
            self._add_formula_cell(ws, row, col, formula, '#,##0')
        return ws

    def build_other_operating_sheet(self):
        ws = self.wb.create_sheet("Other_Operating")
        self._setup_sheet(ws, "Other_Operating")
        row = 1
        row = self._add_title(ws, row, "Other Operating Items")
        ws.cell(row=row, column=1, value="(Placeholder for other operating items)")
        return ws

    def build_taxes_sheet(self):
        ws = self.wb.create_sheet("Taxes")
        self._setup_sheet(ws, "Taxes")
        row = 1
        row = self._add_title(ws, row, "Taxes")
        years = ["Base", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
        for i, year in enumerate(years):
            cell = ws.cell(row=row, column=i+2, value=year)
            cell.font = self.STYLES['header_font']
            cell.alignment = Alignment(horizontal='center')
        row += 1
        ws.cell(row=row, column=1, value="EBIT")
        for i in range(6):
            col = i + 2
            formula = f"=EBITDA_Bridge!{get_column_letter(col)}6+D&A!{get_column_letter(col)}5"
            self._add_formula_cell(ws, row, col, formula, '#,##0')
        row += 1
        ws.cell(row=row, column=1, value="Tax Rate")
        for i in range(6):
            col = i + 2
            cell = ws.cell(row=row, column=col, value=f"=Key_Assumptions!{self._ref('tax_rate')}")
            cell.number_format = '0.0%'
        row += 1
        ws.cell(row=row, column=1, value="Taxes")
        for i in range(6):
            col = i + 2
            formula = f"=-MAX(0,{get_column_letter(col)}{row-2})*{get_column_letter(col)}{row-1}"
            self._add_formula_cell(ws, row, col, formula, '#,##0')
        return ws

    def build_unlevered_fcf_sheet(self):
        ws = self.wb.create_sheet("Unlevered_FCF")
        self._setup_sheet(ws, "Unlevered_FCF")
        row = 1
        row = self._add_title(ws, row, "Unlevered Free Cash Flow")

        years = ["Base", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
        for i, year in enumerate(years):
            cell = ws.cell(row=row, column=i+2, value=year)
            cell.font = self.STYLES['header_font']
            cell.alignment = Alignment(horizontal='center')
        row += 1

        ws.cell(row=row, column=1, value="EBIT")
        for i in range(6):
            col = i + 2
            formula = f"=Taxes!{get_column_letter(col)}4"
            self._add_formula_cell(ws, row, col, formula, '#,##0')
        row += 1

        ws.cell(row=row, column=1, value="Less: Taxes")
        for i in range(6):
            col = i + 2
            formula = f"=Taxes!{get_column_letter(col)}6"
            self._add_formula_cell(ws, row, col, formula, '#,##0')
        row += 1

        ws.cell(row=row, column=1, value="NOPAT")
        for i in range(6):
            col = i + 2
            formula = f"={get_column_letter(col)}{row-2}+{get_column_letter(col)}{row-1}"
            self._add_formula_cell(ws, row, col, formula, '#,##0')
        row += 1

        ws.cell(row=row, column=1, value="Plus: D&A")
        for i in range(6):
            col = i + 2
            formula = f"=-D&A!{get_column_letter(col)}3"
            self._add_formula_cell(ws, row, col, formula, '#,##0')
        row += 1

        ws.cell(row=row, column=1, value="Less: Capex")
        for i in range(6):
            col = i + 2
            formula = f"=Capex!{get_column_letter(col)}3"
            self._add_formula_cell(ws, row, col, formula, '#,##0')
        row += 1

        ws.cell(row=row, column=1, value="Less: Change in NWC")
        for i in range(6):
            col = i + 2
            formula = f"=Working_Capital!{get_column_letter(col)}3"
            self._add_formula_cell(ws, row, col, formula, '#,##0')
        row += 1

        ws.cell(row=row, column=1, value="Unlevered FCF")
        ws.cell(row=row, column=1).font = Font(bold=True)
        for i in range(6):
            col = i + 2
            c = get_column_letter(col)
            formula = f"={c}{row-4}+{c}{row-3}+{c}{row-2}+{c}{row-1}"
            cell = self._add_formula_cell(ws, row, col, formula, '#,##0', bold=True)
            cell.border = self.STYLES['border_bottom']
        return ws

    def build_debt_schedule_sheet(self):
        ws = self.wb.create_sheet("Debt_Schedule")
        self._setup_sheet(ws, "Debt_Schedule")
        row = 1
        row = self._add_title(ws, row, "Debt Schedule")
        ws.cell(row=row, column=1, value="(Placeholder for debt schedule)")
        return ws

    def build_interest_expense_sheet(self):
        ws = self.wb.create_sheet("Interest_Expense")
        self._setup_sheet(ws, "Interest_Expense")
        row = 1
        row = self._add_title(ws, row, "Interest Expense")
        ws.cell(row=row, column=1, value="(Placeholder for interest schedule)")
        return ws

    def build_share_count_sheet(self):
        ws = self.wb.create_sheet("Share_Count")
        self._setup_sheet(ws, "Share_Count")
        row = 1
        row = self._add_title(ws, row, "Share Count")
        ws.cell(row=row, column=1, value="(Placeholder for share count schedule)")
        return ws

    def build_wacc_sheet(self):
        """Build the WACC tab"""
        ws = self.wb.create_sheet("WACC")
        self._setup_sheet(ws, "WACC")
        row = 1
        row = self._add_title(ws, row, "WACC Build")

        row = self._add_section_header(ws, row, "WACC BUILDUP", 4)
        wacc_inputs = [
            ("Risk-Free Rate", self.a.risk_free_rate, '0.00%', 'rf_rate'),
            ("Equity Risk Premium", self.a.equity_risk_premium, '0.00%', 'erp'),
            ("Unlevered Beta", self._unlevered_beta, '0.00', 'unlevered_beta'),
            ("Levered Beta", self._levered_beta, '0.00', 'beta'),
            ("Size Premium", self._size_premium, '0.00%', 'size_prem'),
            ("Company Risk Premium", self._company_specific_risk, '0.00%', 'co_risk'),
            ("Pre-Tax Cost of Debt", self._cost_of_debt, '0.00%', 'cost_debt'),
            ("Debt / Total Capital", self._debt_to_capital, '0.0%', 'debt_cap'),
        ]
        for label, value, fmt, name in wacc_inputs:
            ws.cell(row=row, column=1, value=label)
            self._add_input_cell(ws, row, 2, value, fmt, name)
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="Cost of Equity (CAPM + Adj)")
        cost_equity_formula = f"={self._ref('rf_rate')}+{self._ref('beta')}*{self._ref('erp')}+{self._ref('size_prem')}+{self._ref('co_risk')}"
        self._add_formula_cell(ws, row, 2, cost_equity_formula, '0.00%')
        self._cell_refs['cost_equity'] = f"B{row}"
        row += 1

        ws.cell(row=row, column=1, value="After-Tax Cost of Debt")
        atax_debt_formula = f"={self._ref('cost_debt')}*(1-Key_Assumptions!{self._ref('tax_rate')})"
        self._add_formula_cell(ws, row, 2, atax_debt_formula, '0.00%')
        self._cell_refs['atax_cost_debt'] = f"B{row}"
        row += 1

        ws.cell(row=row, column=1, value="WACC")
        ws.cell(row=row, column=1).font = Font(bold=True)
        wacc_formula = f"={self._ref('cost_equity')}*(1-{self._ref('debt_cap')})+{self._ref('atax_cost_debt')}*{self._ref('debt_cap')}"
        cell = self._add_formula_cell(ws, row, 2, wacc_formula, '0.00%', bold=True)
        cell.fill = self.STYLES['output_fill']
        cell.border = self.STYLES['border']
        self._cell_refs['wacc'] = f"B{row}"
        return ws

    def build_terminal_value_sheet(self):
        ws = self.wb.create_sheet("Terminal_Value")
        self._setup_sheet(ws, "Terminal_Value")
        row = 1
        row = self._add_title(ws, row, "Terminal Value")

        ws.cell(row=row, column=1, value="Terminal Value (Gordon Growth)")
        formula = f"=Unlevered_FCF!G10*(1+Key_Assumptions!{self._ref('term_growth')})/(WACC!{self._ref('wacc')}-Key_Assumptions!{self._ref('term_growth')})"
        self._add_formula_cell(ws, row, 2, formula, '#,##0')
        row += 1

        ws.cell(row=row, column=1, value="Terminal Value (Exit Multiple)")
        formula = f"=EBITDA_Bridge!G6*Key_Assumptions!{self._ref('exit_mult')}"
        self._add_formula_cell(ws, row, 2, formula, '#,##0')
        row += 1

        ws.cell(row=row, column=1, value="PV of TV (Gordon)")
        formula = f"=B2/(1+WACC!{self._ref('wacc')})^IF(Key_Assumptions!{self._ref('mid_year')}=1,5,5)"
        self._add_formula_cell(ws, row, 2, formula, '#,##0')
        row += 1

        ws.cell(row=row, column=1, value="PV of TV (Exit Multiple)")
        formula = f"=B3/(1+WACC!{self._ref('wacc')})^IF(Key_Assumptions!{self._ref('mid_year')}=1,5,5)"
        self._add_formula_cell(ws, row, 2, formula, '#,##0')
        return ws

    def build_dcf_valuation_sheet(self):
        """Build the DCF Valuation tab"""
        ws = self.wb.create_sheet("DCF_Valuation")
        self._setup_sheet(ws, "DCF_Valuation")

        row = 1
        row = self._add_title(ws, row, "Discounted Cash Flow Valuation")

        years = ["Base", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
        for i, year in enumerate(years):
            cell = ws.cell(row=row, column=i+2, value=year)
            cell.font = self.STYLES['header_font']
            cell.alignment = Alignment(horizontal='center')
        row += 1

        ws.cell(row=row, column=1, value="Unlevered FCF")
        for i in range(6):
            col = i + 2
            formula = f"=Unlevered_FCF!{get_column_letter(col)}10"
            self._add_formula_cell(ws, row, col, formula, '#,##0')
        row += 2

        # Discounting Section
        row = self._add_section_header(ws, row, "PRESENT VALUE CALCULATION", 8)

        # Discount Period (with mid-year convention)
        ws.cell(row=row, column=1, value="Discount Period")
        for i in range(5):
            col = i + 3
            # Mid-year: 0.5, 1.5, 2.5... End-year: 1, 2, 3...
            formula = f"=IF(Key_Assumptions!{self._ref('mid_year')}=1,{i}+0.5,{i+1})"
            self._add_formula_cell(ws, row, col, formula, '0.0')
        row += 1

        # Discount Factor
        ws.cell(row=row, column=1, value="Discount Factor")
        for i in range(5):
            col = i + 3
            formula = f"=1/(1+WACC!{self._ref('wacc')})^{get_column_letter(col)}{row-1}"
            self._add_formula_cell(ws, row, col, formula, '0.0000')
        row += 1

        # PV of FCF
        ws.cell(row=row, column=1, value="PV of FCF")
        for i in range(5):
            col = i + 3
            formula = f"={get_column_letter(col)}3*{get_column_letter(col)}{row-1}"
            self._add_formula_cell(ws, row, col, formula, '#,##0')
        row += 2

        # Valuation Summary
        row = self._add_section_header(ws, row, "VALUATION SUMMARY", 4)

        # Sum of PV of FCFs
        ws.cell(row=row, column=1, value="Sum of PV of FCFs")
        formula = f"=SUM(C{row-3}:G{row-3})"
        self._add_formula_cell(ws, row, 2, formula, '#,##0')
        row += 2

        # Gordon Growth Method
        ws.cell(row=row, column=1, value="GORDON GROWTH METHOD")
        ws.cell(row=row, column=1).font = Font(bold=True, italic=True)
        row += 1

        ws.cell(row=row, column=1, value="Enterprise Value")
        formula = f"=B{row-2}+Terminal_Value!B5"
        cell = self._add_formula_cell(ws, row, 2, formula, '#,##0', bold=True)
        row += 1

        ws.cell(row=row, column=1, value="Less: Net Debt")
        formula = f"=-Key_Assumptions!{self._ref('net_debt')}"
        self._add_formula_cell(ws, row, 2, formula, '#,##0')
        row += 1

        ws.cell(row=row, column=1, value="Equity Value")
        formula = f"=B{row-2}+B{row-1}"
        cell = self._add_formula_cell(ws, row, 2, formula, '#,##0', bold=True)
        cell.fill = self.STYLES['output_fill']
        cell.border = self.STYLES['border']
        row += 1

        ws.cell(row=row, column=1, value="Implied Share Price")
        formula = f"=B{row-1}/Key_Assumptions!{self._ref('shares')}"
        cell = self._add_formula_cell(ws, row, 2, formula, '"$"#,##0.00', bold=True)
        cell.fill = self.STYLES['output_fill']
        cell.border = self.STYLES['border']
        cell.font = Font(bold=True, size=12)
        row += 2

        # Exit Multiple Method
        ws.cell(row=row, column=1, value="EXIT MULTIPLE METHOD")
        ws.cell(row=row, column=1).font = Font(bold=True, italic=True)
        row += 1

        ws.cell(row=row, column=1, value="Enterprise Value")
        formula = f"=B{row-7}+Terminal_Value!B6"
        cell = self._add_formula_cell(ws, row, 2, formula, '#,##0', bold=True)
        row += 1

        ws.cell(row=row, column=1, value="Less: Net Debt")
        formula = f"=-Key_Assumptions!{self._ref('net_debt')}"
        self._add_formula_cell(ws, row, 2, formula, '#,##0')
        row += 1

        ws.cell(row=row, column=1, value="Equity Value")
        formula = f"=B{row-2}+B{row-1}"
        cell = self._add_formula_cell(ws, row, 2, formula, '#,##0', bold=True)
        cell.fill = self.STYLES['output_fill']
        cell.border = self.STYLES['border']
        row += 1

        ws.cell(row=row, column=1, value="Implied Share Price")
        formula = f"=B{row-1}/Key_Assumptions!{self._ref('shares')}"
        cell = self._add_formula_cell(ws, row, 2, formula, '"$"#,##0.00', bold=True)
        cell.fill = self.STYLES['output_fill']
        cell.border = self.STYLES['border']
        cell.font = Font(bold=True, size=12)
        return ws

    def build_ev_equity_bridge_sheet(self):
        ws = self.wb.create_sheet("EV_Equity_Bridge")
        self._setup_sheet(ws, "EV_Equity_Bridge")
        row = 1
        row = self._add_title(ws, row, "EV to Equity Bridge")

        ws.cell(row=row, column=1, value="Enterprise Value (Gordon)")
        ws.cell(row=row, column=2, value="=DCF_Valuation!B15").number_format = '#,##0'
        row += 1
        ws.cell(row=row, column=1, value="Less: Net Debt")
        ws.cell(row=row, column=2, value=f"=-Key_Assumptions!{self._ref('net_debt')}").number_format = '#,##0'
        row += 1
        ws.cell(row=row, column=1, value="Equity Value")
        ws.cell(row=row, column=2, value="=B3+B4").number_format = '#,##0'
        row += 1
        ws.cell(row=row, column=1, value="Implied Share Price")
        ws.cell(row=row, column=2, value=f"=B5/Key_Assumptions!{self._ref('shares')}").number_format = '"$"#,##0.00'
        return ws

    def build_sensitivity_sheet(self):
        """Build sensitivity analysis with data tables"""
        ws = self.wb.create_sheet("Sensitivity")
        self._setup_sheet(ws, "Sensitivity")

        row = 1
        ws.cell(row=row, column=1, value="Sensitivity Analysis")
        ws.cell(row=row, column=1).font = self.STYLES['title_font']
        row += 2

        # WACC vs Terminal Growth (Gordon Growth)
        row = self._add_section_header(ws, row, "SHARE PRICE: WACC vs TERMINAL GROWTH (Gordon Growth)", 9)
        row += 1

        # This would need VBA or manual data table setup
        # For now, we'll create the structure
        ws.cell(row=row, column=1, value="Terminal Growth →")
        ws.cell(row=row, column=2, value="WACC ↓")

        # Terminal growth values across top
        tg_values = [0.015, 0.020, 0.025, 0.030, 0.035]
        for i, tg in enumerate(tg_values):
            cell = ws.cell(row=row, column=i+3, value=tg)
            cell.number_format = '0.0%'
            cell.alignment = Alignment(horizontal='center')
        row += 1

        # WACC values down left
        wacc_values = [0.08, 0.09, 0.10, 0.11, 0.12]
        for wacc in wacc_values:
            cell = ws.cell(row=row, column=2, value=wacc)
            cell.number_format = '0.0%'

            # Calculate implied share prices for each combination
            for i, tg in enumerate(tg_values):
                # This is a simplified calculation - real model would reference DCF sheet
                # For MVP, showing structure
                ws.cell(row=row, column=i+3, value="—")
            row += 1

        row += 2
        ws.cell(row=row, column=1, value="Note: Populate data table manually or via VBA")
        ws.cell(row=row, column=1).font = Font(italic=True, color="808080")
        row += 2

        # WACC vs Exit Multiple
        row = self._add_section_header(ws, row, "SHARE PRICE: WACC vs EXIT MULTIPLE", 9)
        row += 1

        ws.cell(row=row, column=1, value="Exit Multiple →")
        ws.cell(row=row, column=2, value="WACC ↓")

        mult_values = [8.0, 9.0, 10.0, 11.0, 12.0]
        for i, mult in enumerate(mult_values):
            cell = ws.cell(row=row, column=i+3, value=mult)
            cell.number_format = '0.0x'
            cell.alignment = Alignment(horizontal='center')
        row += 1

        for wacc in wacc_values:
            cell = ws.cell(row=row, column=2, value=wacc)
            cell.number_format = '0.0%'
            for i in range(5):
                ws.cell(row=row, column=i+3, value="—")
            row += 1

        return ws

    def build_scenario_manager_sheet(self):
        ws = self.wb.create_sheet("Scenario_Manager")
        self._setup_sheet(ws, "Scenario_Manager")
        row = 1
        row = self._add_title(ws, row, "Scenario Manager")

        row = self._add_section_header(ws, row, "DEAL CASE (REALISTIC CABLE TARGET)", 5)
        ws.cell(row=row, column=1, value="Buyer")
        ws.cell(row=row, column=2, value="Comcast (CMCSA)")
        row += 1
        ws.cell(row=row, column=1, value="Target")
        ws.cell(row=row, column=2, value="Altice USA (ATUS)")
        row += 1
        ws.cell(row=row, column=1, value="Deal Type")
        ws.cell(row=row, column=2, value="Strategic Acquisition")
        row += 1
        ws.cell(row=row, column=1, value="Purchase Premium")
        self._add_input_cell(ws, row, 2, 0.30, '0.0%')
        row += 1
        ws.cell(row=row, column=1, value="LTM EV/EBITDA Multiple")
        self._add_input_cell(ws, row, 2, 7.5, '0.0x')
        row += 1
        ws.cell(row=row, column=1, value="Synergies (Run-Rate, $M)")
        self._add_input_cell(ws, row, 2, 300, '#,##0')
        row += 1
        ws.cell(row=row, column=1, value="Funding Mix (Debt / Equity)")
        ws.cell(row=row, column=2, value="60% / 40%")

        return ws

    def build_kpi_dashboard_sheet(self):
        ws = self.wb.create_sheet("KPI_Dashboard")
        self._setup_sheet(ws, "KPI_Dashboard")
        row = 1
        row = self._add_title(ws, row, "KPI Dashboard")
        ws.cell(row=row, column=1, value="Implied Share Price (Gordon)")
        ws.cell(row=row, column=2, value="=DCF_Valuation!B18").number_format = '"$"#,##0.00'
        row += 1
        ws.cell(row=row, column=1, value="Implied Share Price (Exit)")
        ws.cell(row=row, column=2, value="=DCF_Valuation!B24").number_format = '"$"#,##0.00'
        return ws

    def build_trading_comps_sheet(self):
        ws = self.wb.create_sheet("Trading_Comps")
        self._setup_sheet(ws, "Trading_Comps")
        row = 1
        row = self._add_title(ws, row, "Trading Comps")
        ws.cell(row=row, column=1, value="(Placeholder for trading comps)")
        return ws

    def build_transactions_comps_sheet(self):
        ws = self.wb.create_sheet("Transactions_Comps")
        self._setup_sheet(ws, "Transactions_Comps")
        row = 1
        row = self._add_title(ws, row, "Transactions Comps")
        ws.cell(row=row, column=1, value="(Placeholder for transactions comps)")
        return ws

    def build_charts_checks_sheet(self):
        """Build error checking dashboard"""
        ws = self.wb.create_sheet("Charts_Checks")
        self._setup_sheet(ws, "Charts_Checks")

        row = 1
        row = self._add_title(ws, row, "Charts & Model Integrity Checks")

        checks = [
            ("WACC > Terminal Growth", f"=IF(WACC!{self._ref('wacc')}>Key_Assumptions!{self._ref('term_growth')},\"PASS\",\"FAIL\")", "Terminal growth must be less than WACC"),
            ("WACC is Reasonable (5-15%)", f"=IF(AND(WACC!{self._ref('wacc')}>=0.05,WACC!{self._ref('wacc')}<=0.15),\"PASS\",\"WARNING\")", "WACC typically ranges 5-15%"),
            ("Terminal Growth ≤ 4%", f"=IF(Key_Assumptions!{self._ref('term_growth')}<=0.04,\"PASS\",\"WARNING\")", "Should not exceed long-term GDP growth"),
            ("Positive Base Revenue", f"=IF(Key_Assumptions!{self._ref('base_revenue')}>0,\"PASS\",\"FAIL\")", "Revenue must be positive"),
            ("Shares Outstanding > 0", f"=IF(Key_Assumptions!{self._ref('shares')}>0,\"PASS\",\"FAIL\")", "Shares must be positive"),
        ]

        row = self._add_section_header(ws, row, "VALIDATION CHECKS", 5)

        ws.cell(row=row, column=1, value="Check")
        ws.cell(row=row, column=2, value="Status")
        ws.cell(row=row, column=3, value="Description")
        ws.cell(row=row, column=1).font = self.STYLES['header_font']
        ws.cell(row=row, column=2).font = self.STYLES['header_font']
        ws.cell(row=row, column=3).font = self.STYLES['header_font']
        row += 1

        for check_name, formula, description in checks:
            ws.cell(row=row, column=1, value=check_name)
            cell = ws.cell(row=row, column=2, value=formula)
            cell.border = self.STYLES['border']
            ws.cell(row=row, column=3, value=description)
            row += 1

        return ws

    def generate(self, output_path: str):
        """Generate the complete DCF model"""
        self.build_cover_sheet()
        self.build_contents_sheet()
        self.build_key_assumptions_sheet()
        self.build_historical_sheets()
        self.build_revenue_build_sheet()
        self.build_cogs_gross_margin_sheet()
        self.build_opex_sheet()
        self.build_ebitda_bridge_sheet()
        self.build_da_sheet()
        self.build_capex_sheet()
        self.build_working_capital_sheet()
        self.build_other_operating_sheet()
        self.build_taxes_sheet()
        self.build_unlevered_fcf_sheet()
        self.build_debt_schedule_sheet()
        self.build_interest_expense_sheet()
        self.build_share_count_sheet()
        self.build_wacc_sheet()
        self.build_inputs_index_sheet()
        self.build_dcf_valuation_sheet()
        self.build_terminal_value_sheet()
        self.build_ev_equity_bridge_sheet()
        self.build_sensitivity_sheet()
        self.build_scenario_manager_sheet()
        self.build_kpi_dashboard_sheet()
        self.build_trading_comps_sheet()
        self.build_transactions_comps_sheet()
        self.build_charts_checks_sheet()

        self.wb._sheets.sort(key=lambda s: self.TAB_ORDER.index(s.title))

        # Set Cover as the active sheet
        self.wb.active = self.wb["Cover"]

        self.wb.save(output_path)
        return output_path


def generate_dcf_model(assumptions: DCFAssumptions, output_path: str) -> str:
    """Convenience function to generate DCF model"""
    model = ProfessionalDCFModel(assumptions)
    return model.generate(output_path)


if __name__ == "__main__":
    # Test with sample assumptions
    assumptions = DCFAssumptions(
        company_name="Test Corp",
        ticker="TEST",
        base_revenue=1000,
        revenue_growth_y1=0.12,
        revenue_growth_y2=0.10,
        revenue_growth_y3=0.08,
        revenue_growth_y4=0.06,
        revenue_growth_y5=0.04,
        ebitda_margin=0.25,
        da_pct_revenue=0.03,
        capex_pct_revenue=0.04,
        nwc_pct_revenue=0.10,
        tax_rate=0.25,
        risk_free_rate=0.045,
        equity_risk_premium=0.055,
        beta=1.1,
        cost_of_debt=0.06,
        debt_to_capital=0.25,
        terminal_growth=0.025,
        exit_multiple=10.0,
        net_debt=200,
        shares_outstanding=100
    )

    output = generate_dcf_model(assumptions, "test_professional_dcf.xlsx")
    print(f"Generated: {output}")
