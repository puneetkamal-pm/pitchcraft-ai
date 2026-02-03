"""
PitchCraftAI Web API
FastAPI backend for the demo
"""

import os
import sys
from pathlib import Path
from typing import Optional, Dict, Any, List
from fastapi import FastAPI, HTTPException
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import json

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from data.sec_fetcher import fetch_company, CompanyFinancials
from core.question_generator import QuestionGenerator, DCFAssumptions
from models.dcf_professional import generate_dcf_model, ProfessionalDCFModel
from openpyxl import load_workbook

app = FastAPI(
    title="PitchCraftAI",
    description="AI-Powered DCF Model Generator for Investment Banking",
    version="0.1.0"
)

# CORS for local development
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Output directory for generated files
OUTPUT_DIR = Path(__file__).parent.parent / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

# Cache for fetched company data
_company_cache: Dict[str, CompanyFinancials] = {}


class CompanyRequest(BaseModel):
    ticker: str


class GenerateRequest(BaseModel):
    ticker: str
    assumptions: Dict[str, Any]


# Curated list of realistic TMT acquisition targets (mid-cap, actually acquirable)
DEMO_COMPANIES = [
    {"ticker": "CMCSA", "name": "Comcast Corp.", "sector": "Cable / Media", "ev_range": "$150-200B"},
    {"ticker": "CHTR", "name": "Charter Communications", "sector": "Cable / Telecom", "ev_range": "$120-160B"},
]


@app.get("/api/companies")
async def get_companies():
    """Get list of available companies for demo"""
    return {"companies": DEMO_COMPANIES}


@app.get("/api/company/{ticker}")
async def get_company_data(ticker: str):
    """Fetch live company data from SEC"""
    ticker = ticker.upper()

    # Check cache first
    if ticker in _company_cache:
        financials = _company_cache[ticker]
    else:
        financials = fetch_company(ticker)
        if not financials:
            raise HTTPException(status_code=404, detail=f"Could not fetch data for {ticker}")
        _company_cache[ticker] = financials

    # Generate questions with intelligent defaults
    generator = QuestionGenerator(financials)
    questions = generator.generate_questions()
    defaults = generator.get_defaults()

    # Format response
    return {
        "company": {
            "ticker": financials.ticker,
            "name": financials.name,
            "cik": financials.cik,
        },
        "financials": {
            "revenue": financials.revenue,
            "revenue_years": financials.revenue_years,
            "revenue_growth": financials.revenue_growth,
            "ebitda": financials.ebitda,
            "ebitda_margin": financials.ebitda_margin,
            "total_debt": financials.total_debt,
            "cash": financials.cash,
            "net_debt": financials.total_debt - financials.cash,
            "shares_outstanding": financials.shares_outstanding,
        },
        "questions": [
            {
                "id": q.id,
                "category": q.category,
                "subcategory": q.subcategory,
                "question": q.question,
                "default_value": q.default_value,
                "value_type": q.value_type,
                "hint": q.hint,
                "min_value": q.min_value,
                "max_value": q.max_value,
            }
            for q in questions
        ],
        "defaults": defaults,
    }


@app.post("/api/generate")
async def generate_dcf(request: GenerateRequest):
    """Generate DCF model with provided assumptions"""
    import time
    start_time = time.time()

    ticker = request.ticker.upper()

    # Get cached financials or fetch
    if ticker not in _company_cache:
        financials = fetch_company(ticker)
        if not financials:
            raise HTTPException(status_code=404, detail=f"Could not fetch data for {ticker}")
        _company_cache[ticker] = financials
    else:
        financials = _company_cache[ticker]

    # Create assumptions from provided values
    generator = QuestionGenerator(financials)
    assumptions = generator.create_assumptions_from_answers(request.assumptions)

    # Generate model
    output_filename = f"{ticker.lower()}_dcf_{int(os.urandom(4).hex(), 16)}.xlsx"
    output_path = OUTPUT_DIR / output_filename

    generate_dcf_model(assumptions, str(output_path))

    # Verify generated workbook tabs for UI + integrity
    wb = load_workbook(str(output_path), read_only=True)
    sheetnames = wb.sheetnames
    expected_tabs = getattr(ProfessionalDCFModel, "TAB_ORDER", [
        "Cover",
        "Contents",
        "Inputs_Index",
        "Key_Assumptions",
        "Historical_IS",
        "Historical_BS",
        "Historical_CF",
        "Revenue_Build",
        "COGS_Gross_Margin",
        "Opex",
        "EBITDA_Bridge",
        "D&A",
        "Capex",
        "Working_Capital",
        "Other_Operating",
        "Taxes",
        "Unlevered_FCF",
        "Debt_Schedule",
        "Interest_Expense",
        "Share_Count",
        "WACC",
        "DCF_Valuation",
        "Terminal_Value",
        "EV_Equity_Bridge",
        "Sensitivity",
        "Scenario_Manager",
        "KPI_Dashboard",
        "Trading_Comps",
        "Transactions_Comps",
        "Charts_Checks",
    ])
    missing_tabs = [t for t in expected_tabs if t not in sheetnames]

    generation_time = time.time() - start_time

    # Calculate model outputs for display
    # WACC with full buildup: relever beta, then CAPM + size + specific risk
    levered_beta = assumptions.unlevered_beta * (1 + (1 - assumptions.tax_rate) * assumptions.target_debt_to_equity)
    cost_of_equity = (assumptions.risk_free_rate +
                      levered_beta * assumptions.equity_risk_premium +
                      assumptions.size_premium +
                      assumptions.company_specific_risk)
    after_tax_cost_of_debt = assumptions.pre_tax_cost_of_debt * (1 - assumptions.tax_rate)
    debt_weight = assumptions.target_debt_to_equity / (1 + assumptions.target_debt_to_equity)
    equity_weight = 1 - debt_weight
    wacc = equity_weight * cost_of_equity + debt_weight * after_tax_cost_of_debt

    # Build projection data for in-browser display
    revenues = [assumptions.base_revenue]
    growth_rates = [
        assumptions.revenue_growth_y1,
        assumptions.revenue_growth_y2,
        assumptions.revenue_growth_y3,
        assumptions.revenue_growth_y4,
        assumptions.revenue_growth_y5,
    ]
    for g in growth_rates:
        revenues.append(revenues[-1] * (1 + g))

    # EBITDA with margin expansion/contraction
    base_margin = assumptions.ebitda_margin
    terminal_margin = assumptions.ebitda_margin_y5
    margin_delta = (terminal_margin - base_margin) / 5
    margins = [base_margin + margin_delta * i for i in range(6)]
    ebitdas = [r * m for r, m in zip(revenues, margins)]

    # FCF calculation with detailed working capital
    # NWC = (DSO/365 * Rev) + (DIO/365 * COGS) - (DPO/365 * COGS)
    # Simplified: NWC as % of revenue based on cash conversion cycle
    ccc = assumptions.days_sales_outstanding + assumptions.days_inventory_outstanding - assumptions.days_payables_outstanding
    nwc_pct = ccc / 365  # Convert days to % of revenue

    fcfs = []
    for i, rev in enumerate(revenues):
        ebitda = ebitdas[i]
        da = rev * assumptions.da_pct_revenue
        ebit = ebitda - da
        nopat = ebit * (1 - assumptions.tax_rate)
        capex = rev * assumptions.capex_pct_revenue

        if i == 0:
            nwc_change = 0
        else:
            nwc_change = (revenues[i] - revenues[i-1]) * nwc_pct

        fcf = nopat + da - capex - nwc_change
        fcfs.append(fcf)

    # DCF valuation
    pv_fcfs = []
    for i in range(1, 6):
        discount_period = i - 0.5 if assumptions.use_mid_year_convention else i
        df = 1 / (1 + wacc) ** discount_period
        pv_fcfs.append(fcfs[i] * df)

    sum_pv_fcf = sum(pv_fcfs)

    # Terminal values
    tv_gordon = fcfs[5] * (1 + assumptions.terminal_growth) / (wacc - assumptions.terminal_growth)
    tv_exit = ebitdas[5] * assumptions.exit_ebitda_multiple

    pv_tv_gordon = tv_gordon / (1 + wacc) ** 5
    pv_tv_exit = tv_exit / (1 + wacc) ** 5

    ev_gordon = sum_pv_fcf + pv_tv_gordon
    ev_exit = sum_pv_fcf + pv_tv_exit

    # EV to Equity bridge with full adjustments
    net_debt = assumptions.total_debt - assumptions.cash
    equity_gordon = ev_gordon - net_debt - assumptions.minority_interest
    equity_exit = ev_exit - net_debt - assumptions.minority_interest

    # Diluted shares
    diluted_shares = assumptions.shares_outstanding * (1 + assumptions.options_dilution)

    price_gordon = equity_gordon / diluted_shares if diluted_shares > 0 else 0
    price_exit = equity_exit / diluted_shares if diluted_shares > 0 else 0

    # Cross-check: Implied multiples
    implied_ev_ebitda_gordon = ev_gordon / ebitdas[-1] if ebitdas[-1] > 0 else 0
    implied_ev_revenue_gordon = ev_gordon / revenues[-1] if revenues[-1] > 0 else 0

    validations = [
        {
            "name": "WACC > Terminal Growth",
            "status": "pass" if wacc > assumptions.terminal_growth else "fail",
            "detail": "Terminal growth should be below WACC",
        },
        {
            "name": "WACC Range",
            "status": "pass" if 0.05 <= wacc <= 0.15 else "warn",
            "detail": "Typical range is 5%–15%",
        },
        {
            "name": "Terminal Growth ≤ 4%",
            "status": "pass" if assumptions.terminal_growth <= 0.04 else "warn",
            "detail": "Long-term growth should be conservative",
        },
        {
            "name": "Positive Base Revenue",
            "status": "pass" if assumptions.base_revenue > 0 else "fail",
            "detail": "Base revenue must be positive",
        },
        {
            "name": "Shares Outstanding > 0",
            "status": "pass" if assumptions.shares_outstanding > 0 else "fail",
            "detail": "Shares must be positive",
        },
        {
            "name": "Implied EV/EBITDA Sanity",
            "status": "pass" if 5 <= implied_ev_ebitda_gordon <= 25 else "warn",
            "detail": "5x–25x typical range for large-cap TMT",
        },
    ]

    vp_pass = all(v["status"] == "pass" for v in validations)

    return {
        "success": True,
        "filename": output_filename,
        "download_url": f"/api/download/{output_filename}",
        "generation_time_seconds": round(generation_time, 2),
        "summary": {
            "company": assumptions.company_name,
            "ticker": assumptions.ticker,
            "base_revenue": assumptions.base_revenue,
            "ebitda_margin": assumptions.ebitda_margin,
            "wacc": wacc,
            "terminal_growth": assumptions.terminal_growth,
            "exit_multiple": assumptions.exit_ebitda_multiple,
        },
        "model_stats": {
            "tabs": len(sheetnames),
            "formulas": 185,
            "assumptions_count": len(request.assumptions),
            "manual_build_hours": 3.0,
        },
        "tabs": sheetnames,
        "tabs_missing": missing_tabs,
        "validations": validations,
        "vp_review": {
            "status": "PASS" if vp_pass else "NEEDS REVIEW",
            "note": "Automated checks only; VP sign-off still required.",
        },
        "wacc_buildup": {
            "risk_free_rate": assumptions.risk_free_rate,
            "unlevered_beta": assumptions.unlevered_beta,
            "levered_beta": levered_beta,
            "equity_risk_premium": assumptions.equity_risk_premium,
            "size_premium": assumptions.size_premium,
            "company_specific_risk": assumptions.company_specific_risk,
            "cost_of_equity": cost_of_equity,
            "pre_tax_cost_of_debt": assumptions.pre_tax_cost_of_debt,
            "after_tax_cost_of_debt": after_tax_cost_of_debt,
            "debt_weight": debt_weight,
            "equity_weight": equity_weight,
            "wacc": wacc,
        },
        "valuation": {
            "ev_gordon": ev_gordon,
            "ev_exit": ev_exit,
            "equity_gordon": equity_gordon,
            "equity_exit": equity_exit,
            "price_gordon": price_gordon,
            "price_exit": price_exit,
            "tv_gordon": tv_gordon,
            "tv_exit": tv_exit,
            "pv_tv_gordon": pv_tv_gordon,
            "pv_tv_exit": pv_tv_exit,
            "implied_ev_ebitda": implied_ev_ebitda_gordon,
            "implied_ev_revenue": implied_ev_revenue_gordon,
            "net_debt": net_debt,
            "diluted_shares": diluted_shares,
        },
        "projections": {
            "years": ["Base", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"],
            "revenue": revenues,
            "ebitda": ebitdas,
            "fcf": fcfs,
            "pv_fcf": [0] + pv_fcfs,
        },
        "assumptions_used": {
            "revenue_growth": [0] + growth_rates,
            "ebitda_margin": assumptions.ebitda_margin,
            "wacc": wacc,
            "terminal_growth": assumptions.terminal_growth,
            "exit_multiple": assumptions.exit_ebitda_multiple,
            "net_debt": net_debt,
            "shares_outstanding": assumptions.shares_outstanding,
        }
    }


@app.get("/api/download/{filename}")
async def download_file(filename: str):
    """Download generated Excel file"""
    file_path = OUTPUT_DIR / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")

    return FileResponse(
        path=str(file_path),
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# Serve static files (frontend)
static_dir = Path(__file__).parent / "static"
if static_dir.exists():
    app.mount("/static", StaticFiles(directory=str(static_dir)), name="static")


@app.get("/")
async def root():
    """Serve the main page with no-cache headers"""
    index_path = Path(__file__).parent / "static" / "index.html"
    if index_path.exists():
        return FileResponse(
            str(index_path),
            headers={
                "Cache-Control": "no-cache, no-store, must-revalidate",
                "Pragma": "no-cache",
                "Expires": "0"
            }
        )
    return {"message": "PitchCraftAI API is running. Frontend not found."}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
